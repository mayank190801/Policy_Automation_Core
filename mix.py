#In this I am trying to extract table data from word and putting it in excel

from __future__ import (
    absolute_import, division, print_function, unicode_literals
)

from docx import Document
from docx.document import Document as _Document
from docx.oxml.text.paragraph import CT_P
from docx.oxml.table import CT_Tbl
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph


def iter_block_items(parent):
    """
    Generate a reference to each paragraph and table child within *parent*,
    in document order. Each returned value is an instance of either Table or
    Paragraph. *parent* would most commonly be a reference to a main
    Document object, but also works for a _Cell object, which itself can
    contain paragraphs and tables.
    """
    if isinstance(parent, _Document):
        parent_elm = parent.element.body
        # print(parent_elm.xml)
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        raise ValueError("something's not right")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)

def get_table_below_given_tag(doc, tag):
    """
    This function simply returns us the reference to the table below the tags given to us. 
    First find the paragraph, then find the very next table corresponding to the paragraph
    """
    found = False
    for block in iter_block_items(document):
        # print(block.text if isinstance(block, Paragraph) else '<table>')
        if(isinstance(block, Paragraph)):
            if tag in block.text:
                found = True
        else:
            if(found == True):
                return block
                

document = Document('file.docx')
table = get_table_below_given_tag(document, "search word")

table_data = []
for row in table.rows:
    row_data = []
    for cell in row.cells:
        row_data.append(cell.text)
    table_data.append(row_data)

print(table_data)


#In this I am trying to put data into excel file as tables and everything
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
treeData = table_data
ws.title = "genl_Table"

for row in treeData: 
    ws.append(row)

# ft = Font(bold=True)
# for row in ws["A1:C1"]:
#     for cell in row:
#         cell.font = ft

wsnew = wb.create_sheet("genl2_Table")
for row in treeData: 
    wsnew.append(row)


wb.save("MixStore.xlsx")
